import logging
from openpyxl.utils import get_column_letter as col_let
from telegram import Update
from telegram.request import HTTPXRequest
from telegram.ext import ApplicationBuilder, CommandHandler
from datetime import datetime, timedelta
from calendar import monthrange
import os
from dotenv import load_dotenv
import gspread
from gspread import Spreadsheet, Worksheet
from gspread_formatting import set_column_width
from oauth2client.service_account import ServiceAccountCredentials

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.WARN
)


# Headers
sub_headers = ["Приход", "Уход", "Итог"]

# Columns
days_start_col = 3
weekdays = ["Пн ", "Вт ", "Ср ", "Чт ", "Пт ", "Сб ", "Вс "]

# Styles
border_thin = {
    "top": {"style": "SOLID"},
    "bottom": {"style": "SOLID"},
    "left": {"style": "SOLID"},
    "right": {"style": "SOLID"},
}
color_grey = {"red": 0.95, "green": 0.95, "blue": 0.95}
color_white = {"red": 1, "green": 1, "blue": 1}


def format(bold=False, color=None):
    fmt = {
        "textFormat": {"bold": bold},
        "horizontalAlignment": "CENTER",
        "borders": border_thin,
    }
    if color:
        fmt["backgroundColor"] = color

    return fmt


def get_week_start(now: datetime) -> datetime:
    return now - timedelta(days=now.weekday())


def record_attendance(
    action: str, username: str | None, spreadsheet: Spreadsheet
) -> str:
    now = datetime.now()
    sheet = setup_attendance_sheet(spreadsheet, now)
    user_row = setup_user_row(sheet, username, now)

    now = datetime.now()
    col_offset = (now.weekday()) * len(sub_headers)
    col = days_start_col + col_offset
    checkin_cell = (user_row, col)
    checkout_cell = (user_row, col + 1)
    duration_cell = (user_row, col + 2)

    if action == "checkin":
        value = sheet.cell(*checkin_cell).value
        if value:
            return f"🫨 ты уже отметилась сегодня в {value}! пиши Марусе, она поможет тебе с этим"

        checkin_time = now.strftime("%H:%M")
        sheet.update_cell(*checkin_cell, checkin_time)
        return f"🐥 утёнок пришел на работу в {checkin_time}\nхорошей смены!"

    elif action == "checkout":
        checkin_time = sheet.cell(*checkin_cell).value
        if not checkin_time:
            return f"😔 ты забыла отметиться на приход! пожалуйста, сначала отметься через /checkin"

        checkout_time = sheet.cell(*checkout_cell).value
        if checkout_time:
            return f"🤬 ты уже ушла сегодня в {checkout_time}! иди домой"

        checkout_time_str = now.strftime("%H:%M")
        sheet.update_cell(*checkout_cell, checkout_time_str)
        sheet.update_cell(
            *duration_cell,
            f"= {col_let(col+1)}{user_row} - {col_let(col)}{user_row}",
        )

        checkin_time = datetime.strptime(checkin_time, "%H:%M")
        checkout_time = datetime.strptime(checkout_time_str, "%H:%M")
        duration = checkout_time - checkin_time
        duration_hours, remain = divmod(duration.seconds, 3600)
        duration_minutes = remain // 60
        duration_str = f"{duration_hours:02}:{duration_minutes:02}"
        return f"🫡 фух, ушла вовремя! в {checkout_time_str} была сделана вся работа!\nрабочее время: {duration_str}"


def setup_attendance_sheet(spreadsheet: Spreadsheet, now: datetime) -> Worksheet:
    week_start = get_week_start(now)
    week_end = week_start + timedelta(days=6)
    sheet_name = week_start.strftime("%d.%m") + "-" + week_end.strftime("%d.%m")

    try:
        sheet = spreadsheet.worksheet(sheet_name)
    except gspread.exceptions.WorksheetNotFound:
        sheet = spreadsheet.add_worksheet(title=sheet_name, rows="100", cols="50")

        row1 = ["Имя", "Часы"] + [
            f"{day} {(week_start + timedelta(days=i)).strftime('%d.%m')}"
            for i, day in enumerate(weekdays)
            for _ in sub_headers
        ]
        row2 = ["", ""] + sub_headers * len(weekdays)

        col = days_start_col
        formats = [{"range": "A1:B2", "format": format(bold=True)}]
        for i in range(len(weekdays)):
            col_start = col_let(col)
            col_end = col_let(col + len(sub_headers) - 1)
            sheet.merge_cells(f"{col_start}1:{col_end}1")

            color = color_grey if i % 2 == 0 else color_white
            formats.append(
                {
                    "range": f"{col_start}1:{col_end}2",
                    "format": format(bold=True, color=color),
                }
            )
            col += len(sub_headers)

        sheet.update([row1, row2])
        set_column_width(sheet, "C:W", 54)
        sheet.batch_format(formats)

    return sheet


def setup_user_row(sheet: Worksheet, username: str | None, now: datetime) -> int:
    usernames = sheet.col_values(1)
    if username in usernames:
        return usernames.index(username) + 1

    user_row = 3 if len(usernames) == 1 else len(usernames) + 1
    sheet.update_cell(user_row, 1, username)

    formats = [{"range": f"A{user_row}:B{user_row}", "format": format(bold=True)}]
    col = days_start_col
    for i in range(len(weekdays)):
        col_start = col_let(col)
        col_end = col_let(col + len(sub_headers) - 1)
        color = color_grey if i % 2 == 0 else color_white
        formats.append(
            {
                "range": f"{col_start}{user_row}:{col_end}{user_row}",
                "format": format(bold=True, color=color),
            }
        )
        col += len(sub_headers)

    setup_total_formulas(sheet, user_row)
    sheet.batch_format(formats)
    return user_row


def setup_total_formulas(sheet: Worksheet, user_row: int):
    duration_cols_1 = []
    for day in range(1, len(weekdays) + 1):
        duration_col = col_let(days_start_col - 1 + day * len(sub_headers))
        duration_cols_1.append(f"{duration_col}{user_row}")

    sheet.update_acell(f"B{user_row}", f"= {'+'.join(duration_cols_1)}")


class ActionHandler:
    def __init__(self, spreadsheet: Spreadsheet):
        self.spreadsheet = spreadsheet

    async def start(self, update: Update, context):
        await update.message.reply_text(
            "привет, утёнок! теперь мы, прям как на заводе:\n"
            "🐥 /checkin - отметить приход\n"
            "🫡 /checkout - отметить уход"
        )

    async def checkin(self, update: Update, context):
        user = update.effective_user
        response = record_attendance("checkin", user.username, self.spreadsheet)
        await update.message.reply_text(response)

    async def checkout(self, update: Update, context):
        user = update.effective_user
        response = record_attendance("checkout", user.username, self.spreadsheet)
        await update.message.reply_text(response)


def main():
    load_dotenv()

    tg_app = (
        ApplicationBuilder()
        .token(os.getenv("TOKEN"))
        .request(
            HTTPXRequest(
                read_timeout=60.0,
                write_timeout=60.0,
                connect_timeout=60.0,
                pool_timeout=60.0,
            )
        )
        .build()
    )

    scope = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = ServiceAccountCredentials.from_json_keyfile_name(
        os.getenv("SERVICE_ACCOUNT_FILE"), scope
    )
    gspread_client = gspread.authorize(creds)

    spreadsheet = gspread_client.open_by_key(os.getenv("SPREADSHEET_ID"))

    bot = ActionHandler(spreadsheet)

    tg_app.add_handler(CommandHandler("start", bot.start))
    tg_app.add_handler(CommandHandler("checkin", bot.checkin))
    tg_app.add_handler(CommandHandler("checkout", bot.checkout))

    print("Bot is running...")
    tg_app.run_polling()


if __name__ == "__main__":
    main()
